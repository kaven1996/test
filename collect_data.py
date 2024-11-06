
from datetime import datetime
from openpyxl import load_workbook
import os
import pandas as pd
import glob
import re
import json
#set factory different model
class FAB():
    def __init__(self,read_file,export_file):
        self.read_file = read_file
        self.export_file = export_file
        self.model_name = "FAB"
        self.exec_init()
        #month_list = self.get_month_abbr()
        
    def exec_init(self):
        json_data = self.read_json()
        self.rack_fields = json_data.get("FAB",{}).get("rack_key_word")
        self.server_fields = json_data.get("FAB",{}).get("server_key_word")
        self.mb_fields = json_data.get("FAB",{}).get("MB_key_word")
        self.sb_fields = json_data.get("FAB",{}).get("SB_Key_word")
        self.export_file_index = json_data.get("export_loading_index")
        self.get_after_num = json_data.get("get_after_num")
        self.output_sheet_name = json_data.get("output_sheet_name")
        self.model_fac = json_data.get("fac_model")
        self.col = self.get_excel_col()
        self.read_excel()


    def read_excel(self):
        sheet_name = pd.ExcelFile(self.read_file).sheet_names
        for fac in self.model_fac:
            pattern = re.compile(fr'.*{fac}.*',re.IGNORECASE)
            match_sheet = [name for name in sheet_name if pattern.match(name)]
            if not match_sheet[0]:
                print(f"not exit sheet {fac} in {self.read_file} excel")
                continue
            df = pd.read_excel(self.read_file, sheet_name=match_sheet[0])
            last_row = df.tail(1)
            #get columns name list
            columns_index  = df.columns.tolist()
            if fac == "QMF" and self.model_name == "FAB":
                fa_index = columns_index.index(self.rack_fields)
                fa_data = last_row.iloc[:,fa_index:fa_index+self.get_after_num]  
                self.export_to_excel(fa_data,self.export_file_index[fac]["fa"])
                last_row = df.tail(2).head(1)
            rack_index = columns_index.index(self.rack_fields)
            server_index = columns_index.index(self.server_fields)
            mb_index = columns_index.index(self.mb_fields)
            sb_index = columns_index.index(self.sb_fields)
            
            rack_data = last_row.iloc[:,rack_index:rack_index+self.get_after_num]
            server_data = last_row.iloc[:,server_index:server_index+self.get_after_num]
            mb_data = last_row.iloc[:,mb_index:mb_index+self.get_after_num]
            sb_data = last_row.iloc[:,sb_index:sb_index+self.get_after_num]
            
            self.export_to_excel(rack_data,self.export_file_index[fac]["rack"])
            self.export_to_excel(server_data,self.export_file_index[fac]["server"])
            self.export_to_excel(mb_data,self.export_file_index[fac]["mb"])
            self.export_to_excel(sb_data,self.export_file_index[fac]["sb"])

    def get_month_abbr(self):
        now = datetime.now()
        month_abbr = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        current_month_index=now.month - 2
        export_month_index=[(current_month_index + i ) % 12 for i in range(self.get_after_num)]
        month_list = [month_abbr[i] for i in export_month_index]
        return month_list
    
    def get_excel_col(self):
        workbook = load_workbook(self.export_file)
        worksheet = workbook[self.output_sheet_name]
        for cell in worksheet.iter_rows(min_row=1,max_row=1,values_only=True):
            for index, value in enumerate(cell):
                if value == self.rack_fields:
                    col = index
            if col is not None:
                break
        return col

    def export_to_excel(self,data,row):
        df = pd.DataFrame(data)
        with pd.ExcelWriter(self.export_file, mode='a', engine='openpyxl',if_sheet_exists='overlay') as writer:
            df.to_excel(writer,sheet_name=self.output_sheet_name,index=False,header=False,startrow=row-1,startcol=self.col)
    
    def read_json(self):
        script_path=os.path.dirname(__file__)
        json_path=os.path.join(script_path,"collect.json")
        try:
            with open(json_path,'r') as file:
                data = json.load(file)
        except json.JSONDecodeError:
            print("JSON file Decoding Error")
        except FileNotFoundError:
            print("can't find JSON file")
        except Exception as e:
            print(f"error: {e}")
        return data

class model_ama():
    pass

class model_msf():
    pass

def main():
    get_cwd = os.getcwd()
    all_files = glob.glob('*.xlsx')
    fab_pattern = re.compile(r'^fab.*\.xlsx$',re.IGNORECASE)
    ama_pattern = re.compile(r'^ama.*\.xlsx$',re.IGNORECASE)
    msf_pattern = re.compile(r'^msf.*\.xlsx$',re.IGNORECASE)
    loading_pattern = re.compile(r'.*loading.*\.xlsx$',re.IGNORECASE)
    fab_file_name = [file for file in all_files if fab_pattern.match(file) and not file.startswith('~')][0]
    ama_file_name = [file for file in all_files if ama_pattern.match(file) and not file.startswith('~')][0]
    #msf_file_name = [file for file in all_files if msf_pattern.match(file) and not file.startswith('~')][0]
    loading_file_name = [file for file in all_files if loading_pattern.match(file) and not file.startswith('~')][0]
    loading_file_path = os.path.join(get_cwd,loading_file_name)
    #fab_data = model_fab(os.path.join(get_cwd,fab_file_name),loading_file_path)
    
    
main()
